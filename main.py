from openpyxl import load_workbook
from openpyxl.workbook.views import BookView

from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt

import re
import sys
import formulas

class TableModel(QtCore.QAbstractTableModel):
    def __init__(self, data):
        super(TableModel, self).__init__()
        self._data = data

    def data(self, index, role):
        if role == Qt.ItemDataRole.DisplayRole:
            # See below for the nested-list data structure.
            # .row() indexes into the outer list,
            # .column() indexes into the sub-list
            return self._data[index.row()][index.column()]

    def rowCount(self, index):
        # The length of the outer list.
        return len(self._data)

    def columnCount(self, index):
        # The following takes the first sub-list, and returns
        # the length (only works if all rows are an equal length)
        return len(self._data[0])

def pyxl_load_workbook(f):
    # wb = 'workbook'
    wb = load_workbook(filename = f, keep_vba=True, rich_text=True)
    row = wb['Sheet1']

    #ws = 'worksheet'
    ws = wb.active

    arr = []

    for row in ws.iter_rows():
        arr_row = []    
        for cell in row:
            #print(cell.value)
            arr_row.append(cell.value)

        arr.append(arr_row)

    return arr

def pyxl_formulas(arr):

    for i, row in enumerate(arr):
        for j, cell in enumerate(row):
            print(i, j)
            if isinstance(cell, str):
                # re.escape() to exclude special regex chars
                if re.search(cell, '='):
                    print("formula at ", i, j)
                else:
                    print(cell)
            else:
                print(cell)

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.table = QtWidgets.QTableView()

        filename = 'Car Price List.xlsx'

        t_data = pyxl_load_workbook(filename)
        f_data = pyxl_formulas(t_data)


        self.model = TableModel(t_data)
        self.table.setModel(self.model)

        self.setCentralWidget(self.table)


app=QtWidgets.QApplication(sys.argv)
window=MainWindow()

window.show()
app.exec()